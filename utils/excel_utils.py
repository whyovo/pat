import os
import time
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox


def select_excel(self):
    path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if path:
        self.excel_path.set(path)
        check_excel_columns(self, path)
        display_excel_info(self, path)
        update_excel_info_label(self, path)
        messagebox.showinfo("提示", "Excel文件选择完成")


def check_excel_columns(self, path):
    try:
        # 获取默认列
        from configs.excel_header_config import get_default_columns, save_custom_columns

        default_columns = get_default_columns()

        # 读取Excel中的实际列
        df = pd.read_excel(path)
        actual_columns = list(df.columns)

        # 检查是否与默认列不同
        if sorted(actual_columns) != sorted(default_columns):
            # 将Excel中的实际列保存为当前使用的表头
            save_custom_columns(actual_columns)

            # 提示用户当前使用的是非默认表头
            messagebox.showinfo(
                "提示", "检测到Excel使用非默认表头，程序已自动适配当前表头结构"
            )

            # 更新程序中的表头配置
            if hasattr(self, "custom_columns"):
                self.custom_columns = actual_columns

            self.status_bar["text"] = "已适配Excel表头结构"
    except Exception as e:
        print(f"检查Excel列时出错: {str(e)}")


def display_excel_info(self, path):
    try:
        # 添加安全检查，确保output_text已存在
        if not hasattr(self, "output_text"):
            messagebox.showerror("错误", "界面组件尚未初始化完成，请稍后再试")
            return

        self.output_text.delete(1.0, tk.END)

        file_size = os.path.getsize(path)
        self.output_text.insert(tk.END, "=== Excel文件信息 ===\n\n", "header")
        self.output_text.insert(
            tk.END, f"文件名: {os.path.basename(path)}\n", "filename"
        )
        self.output_text.insert(tk.END, f"文件路径: {path}\n", "path")
        self.output_text.insert(tk.END, f"文件大小: {file_size/1024:.2f} KB\n", "info")
        self.output_text.insert(
            tk.END,
            f"最后修改时间: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(os.path.getmtime(path)))}\n\n",
            "info",
        )

        try:
            df = pd.read_excel(path)

            self.output_text.insert(tk.END, "=== 表格结构 ===\n", "header")
            self.output_text.insert(tk.END, f"行数: {len(df)}\n", "info")
            self.output_text.insert(tk.END, f"列数: {len(df.columns)}\n", "info")

            self.output_text.insert(tk.END, "=== 列名列表 ===\n", "header")
            for i, col in enumerate(df.columns, 1):
                self.output_text.insert(tk.END, f"{i}. {col}\n", "column")

            required_columns = [
                "论文年份",
                "论文英文引用信息",
                "论文摘要（中文）",
                "研究问题及创新点",
            ]
            if all(col in df.columns for col in required_columns):
                self.output_text.insert(tk.END, "\n=== 表格状态 ===\n", "header")
                analyzed_count = len(
                    df[df["论文摘要（中文）"].notna() & (df["论文摘要（中文）"] != "")]
                )
                self.output_text.insert(
                    tk.END, f"✓ 已包含 {analyzed_count} 篇已分析论文\n", "success"
                )
            else:
                self.output_text.insert(tk.END, "\n=== 表格状态 ===\n", "header")

        except Exception as e:
            self.output_text.insert(
                tk.END, f"\n读取Excel内容时出错: {str(e)}\n", "error"
            )

        self.output_text.tag_configure(
            "header", foreground="#5ba3e0", font=self.fonts["title"]
        )
        self.output_text.tag_configure(
            "filename", foreground="#ffa852", font=("微软雅黑", 11, "bold")
        )
        self.output_text.tag_configure("path", foreground="#e0e0e0")
        self.output_text.tag_configure("info", foreground="#ffffff")
        self.output_text.tag_configure("column", foreground="#8bc34a")
        self.output_text.tag_configure(
            "success", foreground="#8bc34a", font=("微软雅黑", 10, "bold")
        )
        self.output_text.tag_configure(
            "warning", foreground="#ffb74d", font=("微软雅黑", 10, "bold")
        )
        self.output_text.tag_configure("error", foreground="#ef5350")

    except Exception as e:
        self.output_text.delete(1.0, tk.END)
        self.output_text.insert(tk.END, f"读取Excel文件出错: {str(e)}\n", "error")
        self.output_text.tag_configure("error", foreground="#ef5350")


def create_excel(self):
    path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]
    )
    if path:
        # 获取自定义列或使用默认列
        try:
            from configs.excel_header_config import load_custom_columns

            columns = load_custom_columns()
        except Exception:
            columns = [
                "论文年份",
                "论文英文引用信息",
                "论文摘要（中文）",
                "研究问题及创新点",
                "研究意义",
                "研究对象及特点",
                "实验设置",
                "数据分析方法",
                "重要结论",
                "未来研究展望",
            ]

        df = pd.DataFrame(columns=columns)

        save_to_excel_with_format(df, path)
        self.excel_path.set(path)

        display_excel_info(self, path)
        update_excel_info_label(self, path)

        self.status_bar["text"] = f"已创建新Excel文件: {os.path.basename(path)}"


def update_excel_info_label(self, path):
    if hasattr(self, "excel_info_label"):
        filename = os.path.basename(path)
        self.excel_info_label.config(
            text=f"当前文件: {filename}\n路径: {path}", foreground="#e0e0e0"
        )


def save_to_excel_with_format(df, excel_path, append_mode=True):
    try:
        # 获取自定义表头
        try:
            from configs.excel_header_config import load_custom_columns

            expected_columns = load_custom_columns()
            if not expected_columns or len(expected_columns) == 0:
                # 确保至少有一个字段
                expected_columns = ["论文年份"]
        except Exception:
            expected_columns = ["论文年份"]

        # 如果Excel文件存在且有内容，优先使用Excel中的实际表头
        if os.path.exists(excel_path):
            try:
                existing_df = pd.read_excel(excel_path)
                if not existing_df.empty:
                    expected_columns = list(existing_df.columns)
            except Exception:
                pass

        # 确保DF中有所有预期的列
        for col in expected_columns:
            if col not in df.columns:
                df[col] = "未提供相关信息"
            else:
                df[col] = df[col].fillna("未提供相关信息")
                df[col] = df[col].apply(
                    lambda x: "未提供相关信息" if str(x).strip() == "" else x
                )

        # 只保留自定义表头中的列，以确保顺序一致
        # 但首先检查是否有任何匹配的列
        columns_to_keep = [col for col in expected_columns if col in df.columns]
        if not columns_to_keep:
            # 确保至少保留一列
            if "论文年份" in df.columns:
                columns_to_keep = ["论文年份"]
            elif len(df.columns) > 0:
                columns_to_keep = [df.columns[0]]  # 保留第一列
            else:
                # 创建一个最小列
                df["论文年份"] = "未提供相关信息"
                columns_to_keep = ["论文年份"]

        # 重新排列列顺序
        df = df.reindex(columns=columns_to_keep)

        # 当append_mode为False时，直接保存df而不进行合并
        if not append_mode:
            # 清理数据
            for col in df.columns:
                df[col] = df[col].fillna("未提供相关信息")
                df[col] = df[col].apply(
                    lambda x: (
                        "未提供相关信息"
                        if (
                            str(x).lower() in ("nan", "none", "null", "na", "")
                            or str(x).strip() == ""
                        )
                        else str(x)
                    )
                )

            # 保存数据
            df.to_excel(excel_path, index=False)
            print("数据已保存到Excel（覆盖模式）")
            return True  # 保存成功

        # 下面是原有的append模式逻辑
        final_df = None
        if os.path.exists(excel_path):
            try:
                existing_df = pd.read_excel(excel_path)

                for col in expected_columns:
                    if col not in existing_df.columns:
                        existing_df[col] = "未提供相关信息"

                existing_df = existing_df.reindex(columns=expected_columns)

                if not df.empty:
                    final_df = pd.concat([existing_df, df], ignore_index=True)
                else:
                    final_df = existing_df.copy()
            except Exception:
                final_df = df.copy()
        else:
            final_df = df.copy()

        if final_df is None:
            final_df = df.copy()

        for col in final_df.columns:
            final_df[col] = final_df[col].fillna("未提供相关信息")
            final_df[col] = final_df[col].apply(
                lambda x: (
                    "未提供相关信息"
                    if (
                        str(x).lower() in ("nan", "none", "null", "na", "")
                        or str(x).strip() == ""
                    )
                    else str(x)
                )
            )

        # 保存数据
        final_df.to_excel(excel_path, index=False)
        print("数据已保存到Excel（追加模式）")
        return True

    except Exception as e:
        print(f"保存Excel出错: {str(e)}")
        try:
            df = df.fillna("未提供相关信息")
            df.to_excel(excel_path, index=False)
            print("尝试简单保存Excel成功")
            return True
        except Exception as e2:
            print(f"简单保存也失败: {str(e2)}")
            return False


def format_excel_file(excel_path):
    """单独的Excel格式化函数，与数据保存分离"""
    print("正在应用Excel格式...")
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 查看数据是否实际写入了工作簿
    print(f"Excel工作表行数: {ws.max_row}")
    print(f"Excel工作表列数: {ws.max_column}")

    # 设置首行样式
    print("设置表头样式...")
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        # 设置背景色为浅蓝色
        cell.fill = openpyxl.styles.PatternFill(
            start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
        )

    # ===== 增强的自动调整列宽逻辑 =====
    print("智能调整列宽...")
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        column_name = column[0].value

        # 根据不同的列类型设置不同的宽度策略
        if column_name in [
            "论文摘要（中文）",
            "研究问题及创新点",
            "研究意义",
            "研究对象及特点",
            "实验设置",
            "重要结论",
            "未来研究展望",
        ]:
            # 这些列通常包含长文本，设置较宽的固定宽度
            print(f"为长文本列 {column_name} 设置宽度")
            adjusted_width = 50  # 较宽的固定宽度
        else:
            # 其他列根据内容长度动态调整
            for cell in column:
                if cell.value:
                    try:
                        cell_length = len(str(cell.value))
                        # 对于多行内容，考虑行数和单行最大长度
                        if "\n" in str(cell.value):
                            lines = str(cell.value).split("\n")
                            cell_length = max(len(line) for line in lines)

                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass

            # 设置列宽，但确保合理范围
            min_width = 15 if column_name != "论文年份" else 10  # 年份列可以窄一些
            max_width = 30  # 其他列的最大宽度
            adjusted_width = min(max(max_length + 2, min_width), max_width)

        print(f"设置列 {column_letter} ({column_name}) 宽度为 {adjusted_width}")
        ws.column_dimensions[column_letter].width = adjusted_width

    # ===== 改进的行高调整逻辑 =====
    print("智能调整行高...")
    for i, row in enumerate(ws.rows, 1):
        # 跳过表头行，表头已设置
        if i == 1:
            ws.row_dimensions[i].height = 30
            continue

        # 分析该行所有单元格的内容长度和换行数
        max_lines = 1
        has_long_content = False

        for cell in row:
            if cell.value and isinstance(cell.value, str):
                content = str(cell.value)

                # 计算换行符数量
                newlines = content.count("\n")

                # 估算文本所需行数
                if "\n" in content:
                    lines = content.split("\n")
                    # 考虑每行长度超过单元格宽度导致的自动换行
                    for line in lines:
                        # 假设每个字符宽度约为10像素，单元格宽度约为50个字符
                        char_width = 50
                        line_length = len(line)
                        estimated_lines = max(1, line_length / char_width)
                        max_lines = max(max_lines, newlines + 1, estimated_lines)
                else:
                    # 估计可能的自动换行
                    estimated_lines = len(content) / 50

                max_lines = max(max_lines, int(estimated_lines) + 1)  # +1给予额外空间

                # 检查是否包含长内容
                if len(content) > 200:
                    has_long_content = True

        # 基于内容计算适合的行高
        line_height = 25  # 每行估算高度
        if has_long_content:
            # 对特别长的内容，确保至少有足够空间显示
            min_height = 60
        else:
            min_height = 30

        # 计算最终行高，确保至少达到最小高度，且足够显示所有文本
        row_height = max(min_height, line_height * max_lines)
        # 限制最大行高，避免过高
        row_height = min(row_height, 200)

        print(f"设置行 {i} 高度为 {row_height} ({max_lines}行文本)")
        ws.row_dimensions[i].height = row_height

    # ===== 为所有数据单元格设置自动换行和对齐方式 =====
    print("设置单元格文本格式...")
    for row in ws.iter_rows(min_row=2):  # 从第二行开始（跳过表题行）
        for cell in row:
            # 强制设置自动换行和垂直对齐
            cell.alignment = openpyxl.styles.Alignment(
                wrap_text=True,  # 启用自动换行
                vertical="top",  # 顶部对齐
                horizontal="left",  # 左对齐
            )

            # 为长文本添加适当的填充
            if (
                cell.value
                and isinstance(cell.value, str)
                and len(str(cell.value)) > 100
            ):
                # 为长文本单元格设置淡色背景以增强可读性
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
                )

    # 保存格式化后的文件
    wb.save(
        excel_path
    )  # 这行代码导致了重复保存，但我们需要保留它，因为格式化后需要保存
    print(f"Excel已完成格式化: {excel_path}")
    return True
